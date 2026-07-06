from __future__ import annotations

import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from everida.agents.document import parse_document
from everida.schemas import FieldEvidence, ParsedDocument, PipelineResult, ProductConfig, ValidationIssue
from everida.tools.sql import inspect_sql


BENEFIT_NAMES = [
    "成长守护金",
    "大学教育金",
    "深造教育金",
    "有成关爱金",
    "逐梦远航金",
    "满期保险金",
    "身故保险金",
    "投保人意外身故豁免保险费",
]


@dataclass(frozen=True)
class ProductIdentity:
    risk_code: str
    risk_name: str


PRODUCT_IDENTITY_RE = re.compile(
    r"(?P<code>\b\d{6}\b)\s*[-－]\s*(?P<name>[\u4e00-\u9fffA-Za-z0-9]+?保险(?:（[^）]+）)?)"
)

PRODUCT_NAME_RE = re.compile(r"([\u4e00-\u9fffA-Za-z0-9]+?保险(?:（[^）]+）)?)")
PRODUCT_TYPE_CANDIDATES = ["年金保险", "终身寿险", "两全保险", "健康保险", "意外伤害保险", "重大疾病保险"]
BONUS_TYPE_CANDIDATES = ["分红型", "万能型", "投资连结型"]
MVP_SQL_TABLES = [
    "LMRISKAPP",
    "LMRISKENTRYITEM",
    "LMRISKTORISKTYPE",
    "LMEDORCAL",
    "LFRISK",
    "LMRISKPARAMSDEF",
    "LDAUTOAPPROVECONFIG",
    "LAWAGECALELEMENT",
    "EBSPROTOPROCATALOGDETAIL",
    "LMRISKEDORRULE",
    "LMRISKBASEPARA",
    "LMRISKROLE",
    "LMRISKPAY",
    "LMRISKBASEPARARELA",
    "LMRISKDUTY",
    "LMCALMODE",
    "LMRISKEDORITEM",
    "LMEDORWT",
    "LMEDORZT",
    "LMEDORZT1",
    "LMEDORZTDUTY",
    "LMRISKAMNT",
    "LMRISK",
    "LMRISKAMNTRULE",
    "LMDUTY",
    "LMDUTYCTRL",
    "LMDUTYPAYRELA",
    "LMDUTYGETRELA",
    "LMDUTYPAY",
    "LMDUTYGET",
    "LMDUTYGETCLM",
    "LMDUTYGETALIVE",
    "LMRISKAPPDB",
    "LMRISKCOMCTRL",
    "LMLOAN",
    "LMRISKSORT",
    "LMRISKEDORSERVICE",
    "LMRISKTORISK",
]
REQUIRED_TEMPLATE_SHEETS = ["产品基础信息", "责任定义", "给付责任", "缴费计划", "计算规则"]


def parse_product(spec: str | Path) -> ProductConfig:
    parsed = parse_document(spec)
    text = parsed.markdown
    compact = re.sub(r"\s+", "", text)

    identity = _extract_identity(text)
    risk_code = identity.risk_code
    risk_name = identity.risk_name
    short_name = _derive_short_name(risk_name)
    product_type = _first_candidate(compact, PRODUCT_TYPE_CANDIDATES)
    bonus_type = _first_candidate(compact, BONUS_TYPE_CANDIDATES)
    coverage_schemes = _extract_coverage_schemes(text)
    payment_options = _unique_found(compact, ["趸交", "3年", "5年", "10年", "年交"])
    payment_options.extend(_periods_from_table_codes(compact, ["3", "5", "10"]))
    payment_options = _dedupe(payment_options)
    insurance_periods = _unique_found(compact, ["至25岁", "至30岁", "30年"])
    insurance_periods.extend(_age_periods_from_text(compact))
    insurance_periods = _dedupe(insurance_periods)
    benefit_rules = _unique_found(compact, BENEFIT_NAMES)
    underwriting_rules = _extract_rules(compact, ["投保年龄", "最低保额", "1000元", "性别"])
    preservation_rules = _extract_rules(compact, ["退保", "减保", "保全", "现金价值"])
    claim_rules = _extract_rules(compact, ["身故", "理赔", "风险保额"])
    field_evidence = {
        "risk_code": _evidence_for_value(parsed, risk_code, risk_code, 0.92),
        "risk_name": _evidence_for_value(parsed, risk_name, risk_name, 0.9),
        "short_name": _evidence_for_value(parsed, short_name, short_name, 0.82),
        "product_type": _evidence_for_value(parsed, product_type, product_type, 0.86),
        "bonus_type": _evidence_for_value(parsed, bonus_type, bonus_type, 0.86),
        "coverage_schemes": _evidence_for_list(parsed, coverage_schemes, 0.74, ["保障方案"]),
        "payment_options": _evidence_for_list(parsed, payment_options, 0.72, ["缴费期间", "缴费频率", "缴费方式"]),
        "insurance_periods": _evidence_for_list(parsed, insurance_periods, 0.72, ["保险期间", "保障期间"]),
        "benefit_rules": _evidence_for_list(parsed, benefit_rules, 0.78),
        "underwriting_rules": _evidence_for_list(parsed, underwriting_rules, 0.66),
        "preservation_rules": _evidence_for_list(parsed, preservation_rules, 0.62),
        "claim_rules": _evidence_for_list(parsed, claim_rules, 0.66),
    }

    return ProductConfig(
        risk_code=risk_code,
        risk_name=risk_name,
        short_name=short_name,
        product_type=product_type,
        bonus_type=bonus_type,
        coverage_schemes=coverage_schemes,
        payment_options=payment_options,
        insurance_periods=insurance_periods,
        liabilities=benefit_rules,
        benefit_rules=benefit_rules,
        underwriting_rules=underwriting_rules,
        preservation_rules=preservation_rules,
        claim_rules=claim_rules,
        source_refs=[section.source_ref for section in parsed.sections[:10]],
        field_evidence=field_evidence,
    )


def fill_template(product: ProductConfig, template: str | Path, out: str | Path) -> Path:
    output = Path(out)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template, output)

    workbook = load_workbook(output)
    _fill_product_base_info(workbook, product)
    _fill_duty_definitions(workbook, product)
    _fill_benefit_rules(workbook, product)
    _fill_payment_plans(workbook, product)
    if "Everida产品摘要" in workbook.sheetnames:
        del workbook["Everida产品摘要"]
    sheet = workbook.create_sheet("Everida产品摘要")
    rows = [
        ("字段", "值"),
        ("risk_code", product.risk_code),
        ("risk_name", product.risk_name),
        ("short_name", product.short_name),
        ("product_type", product.product_type),
        ("bonus_type", product.bonus_type),
        ("coverage_schemes", "、".join(product.coverage_schemes)),
        ("payment_options", "、".join(product.payment_options)),
        ("insurance_periods", "、".join(product.insurance_periods)),
        ("benefit_rules", "、".join(product.benefit_rules)),
        ("人工确认项", "请业务人员确认模板各 sheet 的字段映射后再用于生产。"),
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(output)
    return output


def _fill_product_base_info(workbook, product: ProductConfig) -> None:
    if "产品基础信息" not in workbook.sheetnames:
        return
    sheet = workbook["产品基础信息"]
    values_by_field = {
        "riskcode": product.risk_code,
        "riskname": product.risk_name,
        "riskshortname": product.short_name,
    }
    for row in sheet.iter_rows(min_row=2):
        field_cell = row[0]
        if field_cell.value not in values_by_field:
            continue
        target_cell = sheet.cell(row=field_cell.row, column=4)
        target_cell.value = values_by_field[field_cell.value]


def _fill_duty_definitions(workbook, product: ProductConfig) -> None:
    if "责任定义" not in workbook.sheetnames:
        return
    sheet = workbook["责任定义"]
    payment_year = _max_year_value(product.payment_options)
    insurance_period, insurance_period_flag = _max_period_value(product.insurance_periods)
    for index, scheme in enumerate(product.coverage_schemes, start=2):
        if index > sheet.max_row:
            break
        sheet.cell(row=index, column=2).value = f"{scheme}责任"
        if payment_year:
            sheet.cell(row=index, column=3).value = payment_year
            sheet.cell(row=index, column=4).value = "Y-年"
        if insurance_period:
            sheet.cell(row=index, column=5).value = insurance_period
            sheet.cell(row=index, column=6).value = insurance_period_flag


def _fill_benefit_rules(workbook, product: ProductConfig) -> None:
    if "给付责任" not in workbook.sheetnames:
        return
    sheet = workbook["给付责任"]
    for index, benefit_name in enumerate(product.benefit_rules, start=2):
        if index > sheet.max_row:
            break
        sheet.cell(row=index, column=3).value = benefit_name


def _fill_payment_plans(workbook, product: ProductConfig) -> None:
    if "缴费计划" not in workbook.sheetnames:
        return
    sheet = workbook["缴费计划"]
    for index, scheme in enumerate(product.coverage_schemes, start=2):
        if index > sheet.max_row:
            break
        sheet.cell(row=index, column=2).value = f"{scheme}责任缴费"


def generate_sql(product: ProductConfig, table_names: list[str] | None = None) -> str:
    tables = _draft_table_names(table_names)
    payload = json.dumps(
        {
            "risk_code": product.risk_code,
            "risk_name": product.risk_name,
            "short_name": product.short_name,
            "product_type": product.product_type,
            "bonus_type": product.bonus_type,
            "coverage_schemes": product.coverage_schemes,
            "payment_options": product.payment_options,
            "insurance_periods": product.insurance_periods,
            "benefit_rules": product.benefit_rules,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    statements = [
        (
            f"-- {table}: draft placeholder generated from structured product config.\n"
            f"INSERT INTO {table} (riskcode, riskname, everida_review_status, everida_payload)\n"
            f"VALUES ('{product.risk_code}', '{_escape_sql(product.risk_name)}', "
            f"'DRAFT_REVIEW_REQUIRED', '{_escape_sql(payload)}');"
        )
        for table in tables
    ]
    return f"""-- Everida generated draft SQL. Human review required before execution.
-- Product: {product.risk_code} {product.risk_name}
-- Scope: {len(tables)} product factory table classes.
-- Note: column mappings are placeholders and must be reviewed before adapting to production DDL.

{chr(10).join(statements)}
"""


def validate_product(product: ProductConfig, template: str | Path, sql: str | Path) -> str:
    workbook = load_workbook(template, read_only=True, data_only=True)
    inventory = inspect_sql(sql)
    sheet_names = list(workbook.sheetnames)
    sql_text = Path(sql).read_text(encoding="utf-8", errors="ignore")
    missing_template_sheets = _missing_template_sheets(sheet_names)
    missing_sql_tables = _missing_sql_tables(inventory.tables)
    missing_sql_benefits = _missing_sql_benefits(product, sql_text)
    issues = collect_validation_issues(
        product,
        sheet_names,
        inventory.product_codes,
        inventory.tables,
        sql_text=sql_text,
    )
    issue_lines = "\n".join(
        f"- **{issue.severity} / {issue.category}**：{issue.message} 建议：{issue.suggestion}"
        for issue in issues
    )
    if not issue_lines:
        issue_lines = "- 未发现阻断性问题。"
    evidence_table = _field_evidence_markdown(product)

    return f"""# Everida 产品一致性校验报告

## 产品

- 产品编码：{product.risk_code}
- 产品名称：{product.risk_name}
- 产品类型：{product.product_type}
- 分红类型：{product.bonus_type}

## 模板检查

- 模板 Sheet 数：{len(workbook.sheetnames)}
- Sheet：{", ".join(workbook.sheetnames[:20])}

## SQL 检查

- SQL 产品编码：{", ".join(inventory.product_codes) or "未识别"}
- SQL 表数量：{len(inventory.tables)}
- SQL 语句数量：{len(inventory.statements)}

## 核心覆盖检查

- 模板核心 Sheet 覆盖：{len(REQUIRED_TEMPLATE_SHEETS) - len(missing_template_sheets)} / {len(REQUIRED_TEMPLATE_SHEETS)}
- SQL 表类覆盖：{len(MVP_SQL_TABLES) - len(missing_sql_tables)} / {len(MVP_SQL_TABLES)}
- SQL 给付责任覆盖：{len(product.benefit_rules) - len(missing_sql_benefits)} / {len(product.benefit_rules)}

## 字段证据链

{evidence_table}

## 人工确认项

{issue_lines}
"""


def collect_validation_issues(
    product: ProductConfig,
    sheet_names: list[str],
    sql_product_codes: list[str],
    sql_tables: list[str],
    sql_text: str = "",
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    if product.risk_code not in sql_product_codes:
        issues.append(
            ValidationIssue(
                severity="warning",
                category="risk_code_missing_in_sql",
                message=f"SQL 中未识别到产品编码 {product.risk_code}",
                source="product_json",
                target="product_sql",
                source_ref="product.risk_code",
                suggestion="请确认 SQL 文件是否为同一产品。",
            )
        )
    if not sheet_names:
        issues.append(
            ValidationIssue(
                severity="error",
                category="template_empty",
                message="配置模板未识别到任何 sheet",
                source="template_xlsx",
                target="filled_xlsx",
                source_ref="xlsx:workbook",
                suggestion="请检查模板文件是否损坏。",
            )
        )
    missing_template_sheets = _missing_template_sheets(sheet_names)
    if missing_template_sheets:
        issues.append(
            ValidationIssue(
                severity="warning",
                category="template_core_sheets_missing",
                message=f"配置模板缺少核心 Sheet：{', '.join(missing_template_sheets)}",
                source="template_xlsx",
                target="filled_xlsx",
                source_ref="xlsx:workbook",
                suggestion="请确认模板版本是否包含 MVP 核心配置页。",
            )
        )
    if not sql_tables:
        issues.append(
            ValidationIssue(
                severity="warning",
                category="sql_tables_missing",
                message="SQL 中未识别到表名",
                source="product_sql",
                target="validate_report",
                source_ref="sql:all",
                suggestion="请人工确认 SQL 方言或脚本结构。",
            )
        )
    missing_sql_tables = _missing_sql_tables(sql_tables)
    if missing_sql_tables:
        issues.append(
            ValidationIssue(
                severity="warning",
                category="sql_table_classes_missing",
                message=f"SQL 缺少 MVP 表类：{', '.join(missing_sql_tables)}",
                source="product_sql",
                target="validate_report",
                source_ref="sql:tables",
                suggestion="请补充对应表类草稿或确认该产品无需配置。",
            )
        )
    missing_sql_benefits = _missing_sql_benefits(product, sql_text)
    if missing_sql_benefits:
        issues.append(
            ValidationIssue(
                severity="warning",
                category="sql_benefit_rules_missing",
                message=f"SQL 中未发现给付责任：{', '.join(missing_sql_benefits)}",
                source="product_json",
                target="product_sql",
                source_ref="product.benefit_rules",
                suggestion="请确认 SQL 给付责任配置是否遗漏。",
            )
        )
    if product.confidence < 0.85:
        issues.append(
            ValidationIssue(
                severity="info",
                category="manual_review_required",
                message="规则抽取为 MVP 启发式结果，需要人工确认",
                source="spec_docx",
                target="product_json",
                source_ref="docx:all",
                suggestion="后续接入模型抽取和字段级置信度校验。",
            )
        )
    return issues


def run_product_pipeline(spec: str | Path, template: str | Path, sql: str | Path, out_dir: str | Path) -> PipelineResult:
    output_dir = Path(out_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    parsed = parse_document(spec)
    product = parse_product(spec)
    sql_inventory = inspect_sql(sql)
    sql_text = Path(sql).read_text(encoding="utf-8", errors="ignore")
    issues = collect_validation_issues(
        product,
        _sheet_names(template),
        sql_inventory.product_codes,
        sql_inventory.tables,
        sql_text=sql_text,
    )

    artifacts = {
        "parsed_document_json": output_dir / "parsed_document.json",
        "parsed_document_markdown": output_dir / "parsed_document.md",
        "product_json": output_dir / "product.json",
        "filled_xlsx": output_dir / "filled.xlsx",
        "generated_sql": output_dir / "generated.sql",
        "sql_inventory_json": output_dir / "sql_inventory.json",
        "validate_report_markdown": output_dir / "validate_report.md",
        "run_manifest_json": output_dir / "run_manifest.json",
    }

    artifacts["parsed_document_json"].write_text(parsed.model_dump_json(indent=2), encoding="utf-8")
    artifacts["parsed_document_markdown"].write_text(parsed.markdown, encoding="utf-8")
    artifacts["product_json"].write_text(product.model_dump_json(indent=2), encoding="utf-8")
    fill_template(product, template, artifacts["filled_xlsx"])
    artifacts["generated_sql"].write_text(generate_sql(product, sql_inventory.tables), encoding="utf-8")
    artifacts["sql_inventory_json"].write_text(sql_inventory.model_dump_json(indent=2), encoding="utf-8")
    artifacts["validate_report_markdown"].write_text(validate_product(product, template, sql), encoding="utf-8")

    result = PipelineResult.completed(output_dir, artifacts, issues)
    manifest = result.model_dump()
    manifest["inputs"] = {"spec": str(spec), "template": str(template), "sql": str(sql)}
    artifacts["run_manifest_json"].write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def write_product_json(product: ProductConfig, out: str | Path) -> Path:
    path = Path(out)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(product.model_dump_json(indent=2), encoding="utf-8")
    return path


def load_product_json(path: str | Path) -> ProductConfig:
    return ProductConfig.model_validate_json(Path(path).read_text(encoding="utf-8"))


def _extract_identity(text: str) -> ProductIdentity:
    match = PRODUCT_IDENTITY_RE.search(text)
    if match:
        return ProductIdentity(
            risk_code=match.group("code"),
            risk_name=match.group("name").strip(),
        )

    code = _first_match(r"\b([1-9]\d{5})\b", text, "")
    name_match = PRODUCT_NAME_RE.search(text)
    return ProductIdentity(
        risk_code=code,
        risk_name=name_match.group(1).strip() if name_match else "",
    )


def _derive_short_name(risk_name: str) -> str:
    if not risk_name:
        return ""
    name = re.sub(r"（[^）]+）", "", risk_name)
    for suffix in PRODUCT_TYPE_CANDIDATES:
        name = name.replace(suffix, "")
    for prefix in ("中邮", "中国人寿", "中国平安", "太平洋", "新华", "泰康"):
        if name.startswith(prefix) and len(name) > len(prefix) + 1:
            name = name[len(prefix) :]
    return name


def _first_candidate(text: str, candidates: list[str]) -> str:
    for candidate in candidates:
        if candidate in text:
            return candidate
    return ""


def _extract_coverage_schemes(text: str) -> list[str]:
    compact = _normalize_text(text)
    schemes = re.findall(r"保障方案[一二三四五六七八九十A-Za-z0-9]+", compact)
    return _dedupe(schemes)


def _first_match(pattern: str, text: str, default: str) -> str:
    match = re.search(pattern, text)
    return match.group(1) if match and match.groups() else (match.group(0) if match else default)


def _unique_found(text: str, candidates: list[str]) -> list[str]:
    return [candidate for candidate in candidates if candidate in text]


def _dedupe(values: list[str]) -> list[str]:
    return list(dict.fromkeys(values))


def _draft_table_names(table_names: list[str] | None = None) -> list[str]:
    source = table_names or MVP_SQL_TABLES
    return _dedupe([table.upper() for table in source])


def _missing_template_sheets(sheet_names: list[str]) -> list[str]:
    return [sheet for sheet in REQUIRED_TEMPLATE_SHEETS if sheet not in sheet_names]


def _missing_sql_tables(sql_tables: list[str]) -> list[str]:
    present = {table.upper() for table in sql_tables}
    return [table for table in MVP_SQL_TABLES if table not in present]


def _missing_sql_benefits(product: ProductConfig, sql_text: str) -> list[str]:
    if not sql_text:
        return []
    return [benefit for benefit in product.benefit_rules if benefit not in sql_text]


def _periods_from_table_codes(text: str, years: list[str]) -> list[str]:
    found: list[str] = []
    for year in years:
        if f"{year}-Y-年" in text or f"{year}Y年" in text:
            found.append(f"{year}年")
    return found


def _age_periods_from_text(text: str) -> list[str]:
    found: list[str] = []
    for age in ("25", "30"):
        if f"至{age}岁" in text or f"{age}周岁" in text:
            found.append(f"至{age}岁")
    if "30-Y-年" in text or "30年" in text:
        found.append("30年")
    return found


def _extract_rules(text: str, keywords: list[str]) -> list[str]:
    return [keyword for keyword in keywords if keyword in text]


def _max_year_value(values: list[str]) -> str:
    years = []
    for value in values:
        match = re.search(r"(\d+)\s*年", value)
        if match:
            years.append(int(match.group(1)))
    return str(max(years)) if years else ""


def _max_period_value(values: list[str]) -> tuple[str, str]:
    fixed_years: list[int] = []
    age_years: list[int] = []
    for value in values:
        fixed_match = re.search(r"^(\d+)\s*年$", value)
        if fixed_match:
            fixed_years.append(int(fixed_match.group(1)))
            continue
        age_match = re.search(r"至\s*(\d+)\s*岁", value)
        if age_match:
            age_years.append(int(age_match.group(1)))
    if fixed_years:
        return str(max(fixed_years)), "Y-年"
    if age_years:
        return str(max(age_years)), "A-岁"
    return "", ""


def _evidence_for_value(parsed: ParsedDocument, value: str, keyword: str, confidence: float) -> FieldEvidence:
    if not value:
        return FieldEvidence(value=value, source_ref="docx:all", confidence=0.0, evidence_text="")
    source_ref, evidence_text = _find_evidence(parsed, [keyword])
    return FieldEvidence(
        value=value,
        source_ref=source_ref,
        confidence=confidence if evidence_text else 0.45,
        evidence_text=evidence_text,
    )


def _evidence_for_list(
    parsed: ParsedDocument,
    values: list[str],
    confidence: float,
    context_keywords: list[str] | None = None,
) -> FieldEvidence:
    if not values:
        return FieldEvidence(value=[], source_ref="docx:all", confidence=0.0, evidence_text="")
    source_ref, evidence_text = _find_evidence(parsed, values, context_keywords or [])
    return FieldEvidence(
        value=values,
        source_ref=source_ref,
        confidence=confidence if evidence_text else 0.45,
        evidence_text=evidence_text,
    )


def _find_evidence(parsed: ParsedDocument, keywords: list[str], context_keywords: list[str] | None = None) -> tuple[str, str]:
    normalized_keywords = [_normalize_text(keyword) for keyword in keywords if keyword]
    normalized_context = [_normalize_text(keyword) for keyword in context_keywords or [] if keyword]

    if normalized_context:
        prioritized = _find_contextual_evidence(parsed, keywords, normalized_keywords, context_keywords or [], normalized_context)
        if prioritized:
            return prioritized

    for keyword, normalized_keyword in zip(keywords, normalized_keywords, strict=False):
        for section in parsed.sections:
            section_text = f"{section.title}\n{section.text}"
            if keyword in section_text:
                return section.source_ref, _snippet(section_text, [keyword])

        for table in parsed.tables:
            table_text = "\n".join(" | ".join(row) for row in table.rows)
            normalized_text = _normalize_text(table_text)
            if normalized_keyword in normalized_text:
                return table.source_ref, _snippet(table_text, [keyword])

        normalized_markdown = _normalize_text(parsed.markdown)
        if normalized_keyword in normalized_markdown:
            return f"{parsed.kind}:markdown", _snippet(parsed.markdown, [keyword])
    return f"{parsed.kind}:all", ""


def _find_contextual_evidence(
    parsed: ParsedDocument,
    keywords: list[str],
    normalized_keywords: list[str],
    context_keywords: list[str],
    normalized_context: list[str],
) -> tuple[str, str] | None:
    for keyword, normalized_keyword in zip(keywords, normalized_keywords, strict=False):
        for table in parsed.tables:
            table_text = "\n".join(" | ".join(row) for row in table.rows)
            normalized_text = _normalize_text(table_text)
            if normalized_keyword in normalized_text and any(context in normalized_text for context in normalized_context):
                return table.source_ref, _snippet(table_text, [*context_keywords, keyword], radius=180)

    for keyword, normalized_keyword in zip(keywords, normalized_keywords, strict=False):
        for section in parsed.sections:
            section_text = f"{section.title}\n{section.text}"
            normalized_text = _normalize_text(section_text)
            if normalized_keyword in normalized_text and any(context in normalized_text for context in normalized_context):
                return section.source_ref, _snippet(section_text, [keyword])
    return None


def _snippet(text: str, keywords: list[str], radius: int = 90) -> str:
    best_keyword = ""
    for keyword in keywords:
        index = text.find(keyword)
        if index != -1:
            best_keyword = keyword
            start = max(index - radius, 0)
            end = min(index + len(keyword) + radius, len(text))
            return re.sub(r"\s+", " ", text[start:end]).strip()

    normalized_text = _normalize_text(text)
    best_index = -1
    for keyword in keywords:
        normalized_keyword = _normalize_text(keyword)
        index = normalized_text.find(normalized_keyword)
        if index != -1:
            best_index = index
            break
    if best_index == -1:
        return text.strip()[: radius * 2]

    plain = re.sub(r"\s+", " ", text).strip()
    plain_index = plain.find(best_keyword)
    if plain_index == -1:
        plain_index = min(best_index, max(len(plain) - 1, 0))
    start = max(plain_index - radius, 0)
    end = min(plain_index + len(best_keyword) + radius, len(plain))
    return plain[start:end]


def _normalize_text(text: str) -> str:
    normalized = re.sub(r"\s+", "", text)
    normalized = re.sub(r"(\d+)-Y-年", r"\1年", normalized)
    return normalized


def _sheet_names(template: str | Path) -> list[str]:
    workbook = load_workbook(template, read_only=True, data_only=True)
    return list(workbook.sheetnames)


def _escape_sql(value: str) -> str:
    return value.replace("'", "''")


def _field_evidence_markdown(product: ProductConfig) -> str:
    fields = [
        ("risk_code", "产品编码"),
        ("risk_name", "产品名称"),
        ("coverage_schemes", "保障方案"),
        ("payment_options", "缴费期间/方式"),
        ("insurance_periods", "保险期间"),
        ("benefit_rules", "给付责任"),
    ]
    rows = ["| 字段 | 值 | 来源 | 置信度 | 证据 |", "| --- | --- | --- | --- | --- |"]
    for field_name, label in fields:
        evidence = product.field_evidence.get(field_name)
        if not evidence:
            rows.append(f"| {label} |  |  | 0.00 |  |")
            continue
        value = "、".join(evidence.value) if isinstance(evidence.value, list) else str(evidence.value)
        rows.append(
            "| "
            + " | ".join(
                [
                    label,
                    _markdown_cell(value),
                    evidence.source_ref,
                    f"{evidence.confidence:.2f}",
                    _markdown_cell(evidence.evidence_text[:120]),
                ]
            )
            + " |"
        )
    return "\n".join(rows)


def _markdown_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")
